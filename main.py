import os
from openpyxl import load_workbook
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, ContextTypes, MessageHandler, filters
from datetime import datetime, date
import psycopg2

class ConversationState:
    WAITING_FOR_FIO = 1
    WAITING_FOR_BIRTHDATE = 2
    WAITING_FOR_GENDER = 3
    WAITING_FOR_ADDRESS = 4
    WAITING_FOR_TELEPHONE = 5
    WAITING_FOR_EMAIL = 6
    WAITING_FOR_MEDICAL_HISTORY = 7
    WAITING_FOR_FAMILY_HISTORY = 8
    WAITING_FOR_ALLERGIES = 9
    WAITING_FOR_LIFESTYLE = 10
    WAITING_FOR_COMPLAINTS = 11

def load_medical_terms():
    terms = {}
    try:
        file_path = os.path.join(os.path.dirname(__file__), "medical_terms.xlsx")
        wb = load_workbook(filename=file_path)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                terms[row[0].lower()] = row[1]
        return terms
    except Exception as e:
        print(f"Ошибка загрузки medical_terms.xlsx: {e}")
        return {}

def load_icd_codes():
    codes = {}
    try:
        file_path = os.path.join(os.path.dirname(__file__), "codes.xlsx")
        wb = load_workbook(filename=file_path)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                codes[row[0].upper()] = row[1]
        return codes
    except Exception as e:
        print(f"Ошибка загрузки codes.xlsx: {e}")
        return {}

def load_disease_data():
    disease_data = []
    try:
        file_path = os.path.join(os.path.dirname(__file__), "razmetka.xlsx")
        wb = load_workbook(filename=file_path)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            disease_data.append({
                "code": row[0],
                "name": row[1],
                "symptoms": row[2].lower() if row[2] else "",
                "diagnostics": row[3],
                "treatment": row[4],
                "related": row[5]
            })
        return disease_data
    except Exception as e:
        print(f"Ошибка загрузки razmetka.xlsx: {e}")
        return []

DISEASE_DATA = load_disease_data()
MEDICAL_TERMS = load_medical_terms()
ICD_CODES = load_icd_codes()

def test_db_connection():
    try:
        conn = psycopg2.connect(
            dbname="postgres",
            user="myuser",
            password="polina",
            host="192.168.0.6",
            port="5432"
        )
        if conn.status == psycopg2.extensions.STATUS_READY:
            print("Успешно подключились к базе данных PostgreSQL!")
        conn.close()
    except Exception as e:
        print(f"Ошибка подключения к БД: {e}")

def insert_diagnosis(patient_id, code, name, description, diagnosis_date, appointment_date, justification, recommendations):
    try:
        conn = psycopg2.connect(
            dbname="postgres",
            user="myuser",
            password="polina",
            host="192.168.0.6",
            port="5432"
        )
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO диагнозы (
                пациент_id, код_мкб_10, название_заболевания,
                описание_диагноза, дата_обращения, дата_постановки,
                обоснование_диагноза, клинические_рекоменлации
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s);
            """,
            (patient_id, code, name, description, diagnosis_date, appointment_date, justification, recommendations)
        )
        conn.commit()
        cursor.close()
        conn.close()
        print(f"Диагноз успешно добавлен для пациента {patient_id}.")
        return True
    except Exception as e:
        print(f"Ошибка при вставке диагноза: {e}")
        return False

#вставка диагоза починить

def insert_patient_basic(fio, birth_date):
    try:
        parts = fio.split()
        if len(parts) != 3:
            print("ФИО должно содержать 3 части")
            return None

        last_name, first_name, middle_name = parts

        conn = psycopg2.connect(
            dbname="postgres",
            user="myuser",
            password="polina",
            host="192.168.0.6",
            port="5432"
        )
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO пациенты (фамилия, имя, отчество, дата_рождения)
            VALUES (%s, %s, %s, %s)
            RETURNING patient_id;
            """,
            (last_name, first_name, middle_name, birth_date)
        )
        patient_id = cursor.fetchone()[0]
        conn.commit()
        cursor.close()
        conn.close()
        print(f"Пациент добавлен в БД с ID {patient_id}.")
        return patient_id

    except Exception as e:
        print(f"Ошибка при вставке пациента: {e}")
        return None

def update_patient_field(patient_id, field_name, field_value):
    try:
        conn = psycopg2.connect(
            dbname="postgres",
            user="myuser",
            password="polina",
            host="192.168.0.6",
            port="5432"
        )
        cursor = conn.cursor()
        query = f"UPDATE пациенты SET {field_name} = %s WHERE patient_id = %s;"
        cursor.execute(query, (field_value, patient_id))
        conn.commit()
        cursor.close()
        conn.close()
        print(f"Обновлено поле {field_name} для пациента {patient_id}.")
        return True
    except Exception as e:
        print(f"Ошибка при обновлении поля {field_name}: {e}")
        return False

def insert_anamnesis(patient_id, medical_history, family_history, allergies, lifestyle, complaints):
    try:
        record_date = date.today()
        conn = psycopg2.connect(
            dbname="postgres",
            user="myuser",
            password="polina",
            host="192.168.0.6",
            port="5432"
        )
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO анамнез (анамнез_id, дата_записи, история_болезни, семейный_анамнез, аллергии, образ_жизни, жалобы)
            VALUES (%s, CURRENT_DATE, %s, %s, %s, %s, %s);
            """,
            (patient_id, medical_history, family_history, allergies, lifestyle, complaints)
        )
        conn.commit()
        cursor.close()
        conn.close()
        print(f"Анамнез сохранён для пациента {patient_id}.")
        return True
    except Exception as e:
        print(f"Ошибка при вставке анамнеза: {e}")
        return False

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.effective_user.first_name or "друг"
    await update.message.reply_text(f"Здравствуйте, {name}! Напишите Ф.И.О. пациента (Фамилия Имя Отчество).")
    context.user_data.clear()
    context.user_data['state'] = ConversationState.WAITING_FOR_FIO

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if 'state' not in context.user_data:
        await update.message.reply_text("Пожалуйста, начните с команды /start")
        return

    current_state = context.user_data['state']
    text = update.message.text.strip()

    # --- Основные данные пациента ---
    if current_state == ConversationState.WAITING_FOR_FIO:
        if len(text.split()) != 3:
            await update.message.reply_text("Пожалуйста, укажите Ф.И.О. полностью (три слова).")
            return
        context.user_data['fio'] = text
        await update.message.reply_text("Укажите дату рождения пациента (ДД.ММ.ГГГГ).")
        context.user_data['state'] = ConversationState.WAITING_FOR_BIRTHDATE

    elif current_state == ConversationState.WAITING_FOR_BIRTHDATE:
        try:
            birth_date = datetime.strptime(text, "%d.%m.%Y").date()
            context.user_data['birth_date'] = birth_date
            patient_id = insert_patient_basic(context.user_data['fio'], birth_date)
            if not patient_id:
                await update.message.reply_text("Ошибка при сохранении данных пациента. Попробуйте снова.")
                context.user_data.clear()
                return
            context.user_data['patient_id'] = patient_id
            await update.message.reply_text("Укажите пол пациента ('Мужской' или 'Женский').")
            context.user_data['state'] = ConversationState.WAITING_FOR_GENDER
        except ValueError:
            await update.message.reply_text("Неверный формат даты. Используйте ДД.ММ.ГГГГ")

    elif current_state == ConversationState.WAITING_FOR_GENDER:
        gender = text.capitalize()
        if gender not in ['Мужской', 'Женский']:
            await update.message.reply_text("Пожалуйста, введите 'Мужской' или 'Женский'.")
            return
        update_patient_field(context.user_data['patient_id'], 'пол', gender)
        context.user_data['gender'] = gender
        await update.message.reply_text("Укажите адрес проживания пациента.")
        context.user_data['state'] = ConversationState.WAITING_FOR_ADDRESS

    elif current_state == ConversationState.WAITING_FOR_ADDRESS:
        update_patient_field(context.user_data['patient_id'], 'адрес', text)
        context.user_data['address'] = text
        await update.message.reply_text("Укажите номер телефона пациента (11 цифр).")
        context.user_data['state'] = ConversationState.WAITING_FOR_TELEPHONE

    elif current_state == ConversationState.WAITING_FOR_TELEPHONE:
        if not (text.isdigit() and len(text) == 11):
            await update.message.reply_text("Номер телефона должен содержать ровно 11 цифр.")
            return
        update_patient_field(context.user_data['patient_id'], 'телефон', text)
        context.user_data['telephone'] = text
        await update.message.reply_text("Укажите электронную почту пациента.")
        context.user_data['state'] = ConversationState.WAITING_FOR_EMAIL

    elif current_state == ConversationState.WAITING_FOR_EMAIL:
        update_patient_field(context.user_data['patient_id'], 'эл_почта', text)
        context.user_data['email'] = text

        # Вывод всех данных пациента
        await update.message.reply_text(
            f"Данные пациента успешно сохранены:\n"
            f"ФИО: {context.user_data['fio']}\n"
            f"Дата рождения: {context.user_data['birth_date'].strftime('%d.%m.%Y')}\n"
            f"Пол: {context.user_data['gender']}\n"
            f"Адрес: {context.user_data['address']}\n"
            f"Телефон: {context.user_data['telephone']}\n"
            f"Email: {context.user_data['email']}\n\n"
            f"Теперь начнём сбор анамнеза.\n"
            f"Пожалуйста, введите историю болезни пациента (до 255 символов)."
        )
        context.user_data['state'] = ConversationState.WAITING_FOR_MEDICAL_HISTORY

    # --- Опрос анамнеза ---
    elif current_state == ConversationState.WAITING_FOR_MEDICAL_HISTORY:
        if len(text) > 255:
            await update.message.reply_text("Длина текста не должна превышать 255 символов. Пожалуйста, сократите сообщение.")
            return
        context.user_data['medical_history'] = text
        await update.message.reply_text("Введите семейный анамнез пациента (до 255 символов).")
        context.user_data['state'] = ConversationState.WAITING_FOR_FAMILY_HISTORY

    elif current_state == ConversationState.WAITING_FOR_FAMILY_HISTORY:
        if len(text) > 255:
            await update.message.reply_text("Длина текста не должна превышать 255 символов. Пожалуйста, сократите сообщение.")
            return
        context.user_data['family_history'] = text
        await update.message.reply_text("Введите информацию об аллергиях пациента (до 255 символов).")
        context.user_data['state'] = ConversationState.WAITING_FOR_ALLERGIES

    elif current_state == ConversationState.WAITING_FOR_ALLERGIES:
        if len(text) > 255:
            await update.message.reply_text("Длина текста не должна превышать 255 символов. Пожалуйста, сократите сообщение.")
            return
        context.user_data['allergies'] = text
        await update.message.reply_text("Введите информацию об образе жизни пациента (до 255 символов).")
        context.user_data['state'] = ConversationState.WAITING_FOR_LIFESTYLE

    elif current_state == ConversationState.WAITING_FOR_LIFESTYLE:
        if len(text) > 255:
            await update.message.reply_text("Длина текста не должна превышать 255 символов. Пожалуйста, сократите сообщение.")
            return
        context.user_data['lifestyle'] = text
        await update.message.reply_text("Введите жалобы пациента на момент обращения (до 255 символов).")
        context.user_data['state'] = ConversationState.WAITING_FOR_COMPLAINTS

    elif current_state == ConversationState.WAITING_FOR_COMPLAINTS:
        if len(text) > 255:
            await update.message.reply_text("Длина текста не должна превышать 255 символов. Пожалуйста, сократите сообщение.")
            return
        context.user_data['complaints'] = text.lower()

        # Сопоставление симптомов
        found_diseases = []
        for disease in DISEASE_DATA:
            if all(symptom.strip() in context.user_data['complaints'] for symptom in disease['symptoms'].split(", ")):
                found_diseases.append(disease)

        # Сохраняем анамнез
        success = insert_anamnesis(
            context.user_data['patient_id'],
            context.user_data.get('medical_history', ''),
            context.user_data.get('family_history', ''),
            context.user_data.get('allergies', ''),
            context.user_data.get('lifestyle', ''),
            context.user_data.get('complaints', '')
        )

        if success:
            await update.message.reply_text("Анамнез пациента успешно сохранён.")
        else:
            await update.message.reply_text("Произошла ошибка при сохранении анамнеза.")

        # Вывод найденных заболеваний
        if found_diseases:
            for disease in found_diseases:
                await update.message.reply_text(
                    f"🔍 Обнаружено соответствие заболеванию:\n"
                    f"📌 Название: {disease['name']}\n"
                   # f"🩺 Симптомы: {disease['symptoms']}\n"
                    f"🧪 Диагностика: {disease['diagnostics']}\n"
                    f"💊 Лечение: {disease['treatment']}\n"
                    f"🔗 Связанные заболевания: {disease['related']}"
                )
        else:
            await update.message.reply_text("❗️ Не удалось найти заболевание по указанным жалобам.")

        context.user_data.clear()


    else:
        await update.message.reply_text("Пожалуйста, начните с команды /start")
        context.user_data.clear()

async def terms_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['state'] = None  # сбрасываем состояние
    if not context.args:
        await update.message.reply_text("✏️ Введите термин после команды. Пример:\n/terms гипертония")
        return

    term = " ".join(context.args).lower()
    if term in MEDICAL_TERMS:
        await update.message.reply_text(f"📖 {term.capitalize()}:\n{MEDICAL_TERMS[term]}")
    else:
        await update.message.reply_text("❌ Термин не найден в справочнике.")

async def codes_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['state'] = None  # сбрасываем состояние
    if not context.args:
        await update.message.reply_text("✏️ Введите код МКБ-10 после команды. Пример:\n/codes J45")
        return

    code = context.args[0].upper()
    if code in ICD_CODES:
        await update.message.reply_text(f"📘 Код {code}:\n{ICD_CODES[code]}")
    else:
        await update.message.reply_text("❌ Код не найден. Проверьте правильность ввода.")

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Доступные команды:\n"
        "/start — начать ввод данных пациента\n"
        "/terms — справочник медицинских терминов\n"
        "/codes — расшифровка кодов МКБ-10\n"
        "/help — список команд"
    )

if __name__ == '__main__':
    TOKEN = '7903839198:AAHD7_C1qic4ic9Xc8ei53XVSOAoOmZ_Bi8'  # замените на ваш токен

    test_db_connection()

    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("terms", terms_command))
    app.add_handler(CommandHandler("codes", codes_command))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(MessageHandler(filters.TEXT & (~filters.COMMAND), handle_message))

    print("Бот запущен...")
    app.run_polling()
